import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from ..database import get_db


# Map internal categories to user-friendly display names
CATEGORY_DISPLAY = {
    "groceries": "Grocery",
    "dining": "Dine Out",
    "gas": "Gas",
    "transportation": "Transportation",
    "auto": "Car Loan",
    "utilities": "Utilities",
    "rent": "Rent",
    "insurance": "Insurance",
    "healthcare": "Health",
    "shopping": "Miscellaneous",
    "subscriptions": "Subscription",
    "travel": "Travel",
    "education": "Education",
    "entertainment": "Entertainment",
    "investments": "Investments",
    "transfer": "Transfer",
    "income": "Income",
    "fees": "Fees",
    "other": "Miscellaneous",
    "uncategorized": "Miscellaneous",
}

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generate_report(
    profile_id: int | None = None,
    year: int | None = None,
    month: int | None = None,
) -> io.BytesIO:
    db = get_db()

    # Build filters
    conditions = []
    params: list = []
    if profile_id:
        conditions.append("t.account_id IN (SELECT id FROM accounts WHERE profile_id = ?)")
        params.append(profile_id)
    if year and month:
        conditions.append("strftime('%Y', t.date) = ?")
        conditions.append("strftime('%m', t.date) = ?")
        params.extend([str(year), f"{month:02d}"])
    elif year:
        conditions.append("strftime('%Y', t.date) = ?")
        params.append(str(year))

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    # Fetch all transactions
    rows = db.execute(
        f"""SELECT t.date, t.description, t.amount, t.tx_type, t.category,
                   a.account_name, a.institution
            FROM transactions t
            JOIN accounts a ON t.account_id = a.id
            {where}
            ORDER BY t.date""",
        params,
    ).fetchall()
    db.close()

    expenses = [r for r in rows if r["tx_type"] == "debit" and r["category"] != "transfer"]
    incomes = [r for r in rows if r["tx_type"] == "credit" and r["category"] != "transfer"]
    transfers = [r for r in rows if r["category"] == "transfer"]

    wb = Workbook()

    # ==================== Sheet 1: Expenses ====================
    ws_exp = wb.active
    ws_exp.title = "Expenses"
    ws_exp.append(["Date", "Type", "Source", "Description", "Amount"])
    _style_header(ws_exp, 5)

    for r in expenses:
        dt = datetime.strptime(r["date"], "%Y-%m-%d") if isinstance(r["date"], str) else r["date"]
        cat_display = CATEGORY_DISPLAY.get(r["category"], r["category"].title())
        ws_exp.append([dt, cat_display, r["description"], "", r["amount"]])

    # Format date and amount columns
    for row_idx in range(2, ws_exp.max_row + 1):
        ws_exp.cell(row=row_idx, column=1).number_format = "YYYY-MM-DD"
        ws_exp.cell(row=row_idx, column=5).number_format = CURRENCY_FMT
        for col in range(1, 6):
            ws_exp.cell(row=row_idx, column=col).border = THIN_BORDER

    # Total row
    total_row = ws_exp.max_row + 1
    ws_exp.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
    ws_exp.cell(row=total_row, column=1).fill = TOTAL_FILL
    total_exp = sum(r["amount"] for r in expenses)
    ws_exp.cell(row=total_row, column=5, value=total_exp).font = TOTAL_FONT
    ws_exp.cell(row=total_row, column=5).fill = TOTAL_FILL
    ws_exp.cell(row=total_row, column=5).number_format = CURRENCY_FMT
    _auto_width(ws_exp)

    # ==================== Sheet 2: Income ====================
    ws_inc = wb.create_sheet("Income")
    ws_inc.append(["Source", "Amount"])
    _style_header(ws_inc, 2)

    # Group income by source description
    income_by_source: dict[str, float] = {}
    for r in incomes:
        source = r["description"]
        income_by_source[source] = income_by_source.get(source, 0) + r["amount"]

    for source, amount in sorted(income_by_source.items(), key=lambda x: -x[1]):
        ws_inc.append([source, amount])

    for row_idx in range(2, ws_inc.max_row + 1):
        ws_inc.cell(row=row_idx, column=2).number_format = CURRENCY_FMT
        for col in range(1, 3):
            ws_inc.cell(row=row_idx, column=col).border = THIN_BORDER

    # Income subtotals by category type (Amazon, Interest, etc.)
    total_row = ws_inc.max_row + 2
    ws_inc.cell(row=total_row, column=1, value="Total ").font = TOTAL_FONT
    ws_inc.cell(row=total_row, column=1).fill = TOTAL_FILL
    total_inc = sum(r["amount"] for r in incomes)
    ws_inc.cell(row=total_row, column=2, value=total_inc).font = TOTAL_FONT
    ws_inc.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws_inc.cell(row=total_row, column=2).number_format = CURRENCY_FMT
    _auto_width(ws_inc)

    # ==================== Sheet 3: Summary_Expenses ====================
    ws_sum_exp = wb.create_sheet("Summary_Expenses")
    ws_sum_exp.append(["Source/Category", "Total Amount"])
    _style_header(ws_sum_exp, 2)

    # Group expenses by display category
    cat_totals: dict[str, float] = {}
    for r in expenses:
        cat_display = CATEGORY_DISPLAY.get(r["category"], r["category"].title())
        cat_totals[cat_display] = cat_totals.get(cat_display, 0) + r["amount"]

    for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1]):
        ws_sum_exp.append([cat, total])

    for row_idx in range(2, ws_sum_exp.max_row + 1):
        ws_sum_exp.cell(row=row_idx, column=2).number_format = CURRENCY_FMT
        for col in range(1, 3):
            ws_sum_exp.cell(row=row_idx, column=col).border = THIN_BORDER

    total_row = ws_sum_exp.max_row + 1
    ws_sum_exp.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
    ws_sum_exp.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws_sum_exp.cell(row=total_row, column=2, value=total_exp).font = TOTAL_FONT
    ws_sum_exp.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws_sum_exp.cell(row=total_row, column=2).number_format = CURRENCY_FMT
    _auto_width(ws_sum_exp)

    # ==================== Sheet 4: Summary_Totals ====================
    ws_totals = wb.create_sheet("Summary_Totals")
    ws_totals.append(["Category", "Amount"])
    _style_header(ws_totals, 2)

    total_transfers = sum(r["amount"] for r in transfers)

    summary_rows = [
        ("Total Income", total_inc),
        ("Total Expenses", total_exp),
        ("Net Savings", total_inc - total_exp),
        ("Internal Transfers", total_transfers),
    ]
    for label, amount in summary_rows:
        ws_totals.append([label, amount])

    for row_idx in range(2, ws_totals.max_row + 1):
        ws_totals.cell(row=row_idx, column=2).number_format = CURRENCY_FMT
        ws_totals.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
        for col in range(1, 3):
            ws_totals.cell(row=row_idx, column=col).border = THIN_BORDER

    # Highlight Net Savings row
    net_row = 4  # row 4 = Net Savings
    ws_totals.cell(row=net_row, column=1).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ws_totals.cell(row=net_row, column=2).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ws_totals.cell(row=net_row, column=2).font = Font(bold=True, size=12, color="16A34A")
    _auto_width(ws_totals)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
